####################################################################################
# Nyolcadikos szóbeli felvételi vizsgabeosztó
# (Első, fapados verzió)
#
# Készítette: Venczel József
#
# Debrecen, 2021. 02. 26.
####################################################################################

from openpyxl import load_workbook
import locale
import sys

# A betöltendő adatokat tartalmazó fájl helye és neve
adatok_tbl = load_workbook("../2020-2021/alapadatok.xlsx")

# Ide kell beírni a táblázatban található, a vizsganapokat jelölő munkalapok neveit.
napok = ["márc.08. (H)", "márc.09. (K)", "márc.10. (Sz)", "márc.11. (Cs)"]















vizsgazok_ml = adatok_tbl["Adatok"]
vizsgazokNapiMax = 5


nap = 0

abc = " -aáäbcdeéfghiíjklmnoóöőpqrstuúüűvwxyzAÁBCDEÉFGHIÍJKLMNOÓÖŐPQRSTUÚÜŰVWXYZ()0123456789"

def strkonv(s):
    kod = []
    for x in s:
        if x in abc:
            kod.append(abc.index(x))
        else:
            print(s)
            sys.exit()
    return kod

class VizsgaIdopont():
    def __init__(self, bizoszlop, biznev, felkido, idopont, nap):
        self.bizoszlop = bizoszlop
        self.biznev = biznev
        self.felkido = felkido
        self.idopont = idopont
        self.nap = nap

    def __eq__(self, masik):
        return (self.bizoszlop == masik.bizoszlop) and (self.idopont == masik.idopont) and (self.nap == masik.nap)

class Diak():
    def __init__(self, om, nev, legkorabban):
        self.om = om
        self.nev = nev
        self.legkorabban = legkorabban
        self.tantargyak = []
        self.vizsgaIdopontok = []

class Diakok():
    def __init__(self):
        self.lista = []

    def ujDiak(self, om, nev, legkorabban):
        self.lista.append(Diak(om, nev, legkorabban))

    def ujVizsgaidopont(self, diakOm, bizoszlop, biznev, felkido, idopont, nap):
        next(x.vizsgaIdopontok for x in self.lista if x.om == diakOm).append(VizsgaIdopont(bizoszlop, biznev, felkido, idopont, nap))

    def nevazonossag(self):
        for x in self.lista:
            azonosak = [y for y in self.lista if y.nev == x.nev]
            if(len(azonosak) > 1):
                for y in azonosak:
                    y.nev = y.nev + " (" + str(y.om)[-4:] + ")"



# Egy-egy tantárgyhoz több bizottság is tartozhat
# A bizottsagot a tantárgyon belül a táblázatbeli oszlopszáma azonosítja
#
class BizHely():
    def __init__(self, bizoszlop, epresz, terem):
        self.bizoszlop = bizoszlop
        self.epresz = epresz
        self.terem = terem


class Bizottsag():
    def __init__(self, tantargy, felkido):
        self.tantargy = tantargy
        self.felkido = felkido
        self.bizhelyek = []


class Bizottsagok():
    def __init__(self):
        self.lista = []

    def ujBizottsag(self, tantargy, bizoszlop, felkido, epresz, terem):
        if(not self.lista or tantargy not in [x.tantargy for x in self.lista]):
            self.lista.append(Bizottsag(tantargy, felkido))
            self.lista[-1].bizhelyek.append(BizHely(bizoszlop, epresz, terem))
        else:
            next(bizottsag.bizhelyek for bizottsag in self.lista if bizottsag.tantargy == tantargy ).append(BizHely(bizoszlop, epresz, terem))

    def vanFelkido(self, bizoszlop):
        tantargy = beosztas_ml[4][bizoszlop].value
        return next(x.felkido for x in self.lista if tantargy == x.tantargy)


# Megkeresi az első rendelkezésre álló vizsgahelyet egy bizottságon belül az "egy bizottságosoknál"
#
def helykereso_1(bizoszlop, minsor):
    sor = minsor+5
    while(beosztas_ml[sor][bizoszlop].value is not None):
        sor+=1
    return sor

# Egy adott bizottságnál, az akthelyből kiindulva megkeresi, hogy az akthely előtt van-e szabad vizsgaidőpont.
#
# def helykereso_elotte(bizottsag, minhely, akthely):
#     sor = akthely
#     idopontok = []
#     while( sor > minhely+3 and ( (type(beosztas_ml[sor][bizottsag].value) == int) or beosztas_ml[sor][bizottsag].value is None) ):
#         if(beosztas_ml[sor][bizottsag].value is None):
#             idopontok.append(sor)
#         sor-=1
#     return idopontok

# # Egy adott bizottságnál, az akthelytől elindulva megkeresi, hogy van-e az akthely után szabad vizsgaidőpont.
# #
# def helykereso_utana(bizottsag, minhely, akthely):
#     sor = akthely if akthely >= minhely else minhely
#     idopontok = []
#     while(beosztas_ml[sor][0].value is not None):
#         if(beosztas_ml[sor][bizottsag].value is None):
    #         idopontok.append(sor)
    #     sor+=1
    # return idopontok


diakok_akt = Diakok()
diakok_holnap = Diakok()
diakok = Diakok()

tantargyak = []
tantargyak_terheltsege = []

#===============================================================================================
# Adatok kigyűjtése, elrendezése
#
#------------------------------------------------------
#

# A tantárgyak kigyűjtése a vizsgázók munkalapjáról.
# Valamint összeszámolja hány vizsga van az egyes tantárgyakból.
#
for tantargy in enumerate(vizsgazok_ml[1]):
    if tantargy[0] > 2:
        tantargyak.append(tantargy[1].value)
        tantargyak_terheltsege.append(0)
        for sor in vizsgazok_ml.rows:
            if(sor[tantargy[0]].value == "x"):
                tantargyak_terheltsege[-1] += 1

print("Feleletek száma tantárgyanként")
for x in enumerate(tantargyak):
    print(f"{x[1]:10}: {tantargyak_terheltsege[x[0]]:4}")

# A vizsgázók kigyűjtése
#
for sor in vizsgazok_ml.rows:
    if(sor[0].value != "OM"):
        sor[0].value = int(sor[0].value)
        diakok_akt.ujDiak( sor[0].value, sor[1].value, int(0 if sor[2].value is None else sor[2].value) )

        for x in enumerate(sor):
            if(x[0]>2 and str(x[1].value) == "x"):
                diakok_akt.lista[-1].tantargyak.append(tantargyak[x[0]-3])

diakok_akt.nevazonossag()

# Vizsgázók sorbarendezése a tantárgyaik száma szerint, csökkenően
#
diakok_akt.lista.sort(key=lambda x: len(x.tantargyak), reverse=True)

diakok.lista = diakok_akt.lista
#////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////
# Egy nap beosztása
#////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////
# Ezt kell ismételgetni, miközben átmegyünk a következő napra
#
print("Vizsgák beosztása...")
tovabb = True

while(tovabb):

    beosztas_ml = adatok_tbl[napok[nap]]

    for col in range(beosztas_ml.max_column):
        if(col > 0):
            beosztas_ml.column_dimensions[chr(65+col)].width = 12

    # A bizottságok csoportosítása tantárgyanként
    # Végigmegy a beosztáson és kigyűjti tantárgyanként a bizottságokat. A bizottságokat a táblázatban elfoglalt helyük,
    # azaz az oszlopsorszámuk szerint különbözteti meg.
    #
    bizottsagok = Bizottsagok()
    for tantargy in enumerate(beosztas_ml[4]):
        if str(tantargy[1].value) != "Tantárgyak:":
            bizottsagok.ujBizottsag(tantargy[1].value, tantargy[0], True if beosztas_ml[5][tantargy[0]].value is not None else False,
                                        beosztas_ml[3][tantargy[0]].value, beosztas_ml[2][tantargy[0]].value)

    # Az Adatok és a Beosztás lapon egyeznek-e a tantárgyak feliratai
    #
    if(len(tantargyak) != len(bizottsagok.lista)):
        print("A tantárgyak nem egyeznek az Adatok és a Beosztások munkalapon!")
    else:
        for x in tantargyak:
            if(x not in [biz.tantargy for biz in bizottsagok.lista]):
                print(f"Nincs {x} nevű tantárgy a bizottságok tantárgyai között!")

    # Meghatározza, hányan vizsgázhatnak egy bizottságnál egy napon
    #
    while(beosztas_ml[vizsgazokNapiMax][0].value is not None):
        vizsgazokNapiMax += 1
    vizsgazokNapiMax -= 1


    # Ez a ciklus osztja be a diákokat egy vizsganapra, illetve teszi át a következőre, ha szükséges
    #
    print(f"{napok[nap]}: ", end="")
    szlo = 0
    for d in diakok_akt.lista:

        if(szlo%int(len(diakok_akt.lista)/10) == 0):
            print(f"{int(szlo/len(diakok_akt.lista)*100):3}%", end = " - ")
        szlo += 1

        # Akik csak 1 osztályba jelentkeztek ---------------------------------------------------------------------------
        #
        if( len(d.tantargyak) == 1 ):

            helyek = []
            for x in next(biz.bizhelyek for biz in bizottsagok.lista if biz.tantargy == d.tantargyak[0]):
                h = helykereso_1(x.bizoszlop, d.legkorabban)
                if(h > -1):
                    helyek.append(VizsgaIdopont(x.bizoszlop, beosztas_ml[1][x.bizoszlop].value, bizottsagok.vanFelkido(x.bizoszlop), h, next(biz.felkido for biz in bizottsagok.lista if biz.tantargy == d.tantargyak[0])))

            leghamarabb = min([a.idopont for a in helyek])
            if(leghamarabb > vizsgazokNapiMax):
                diakok_holnap.lista.append(d)
            else:
                bizoszlop = next(a.bizoszlop for a in helyek if a.idopont == leghamarabb)
                beosztas_ml[ leghamarabb ][ bizoszlop ].value = d.om
                diakok.ujVizsgaidopont(d.om, bizoszlop, beosztas_ml[1][bizoszlop].value, bizottsagok.vanFelkido(bizoszlop),  leghamarabb, napok[nap])

        # Akik 2 osztályba is beadták a jelentkezést -------------------------------------------------------------------
        #
        elif( len(d.tantargyak) == 2 ):
            helykombinaciok_2 = []

            for k in [ [0, 1], [1, 0] ]:

                bizottsag1 = next(biz for biz in bizottsagok.lista if biz.tantargy == d.tantargyak[k[0]])
                bizottsag2 = next(biz for biz in bizottsagok.lista if biz.tantargy == d.tantargyak[k[1]])

                if bizottsag1.felkido or (not bizottsag1.felkido and  not bizottsag2.felkido):

                    set1 = set({})
                    for x in bizottsag1.bizhelyek:
                        set1.add(x.epresz)

                    set2 = set({})
                    for x in bizottsag2.bizhelyek:
                        set2.add(x.epresz)

                    nincs_egy_szinten = False if len(set1 & set2)>0 else True

                    for bizhely1 in bizottsag1.bizhelyek:
                        for bizhely2 in bizottsag2.bizhelyek:
                            if(bizhely1.epresz == bizhely2.epresz or nincs_egy_szinten):
                                h = helykereso_1(bizhely1.bizoszlop, d.legkorabban)
                                utana = h+(5 if bizottsag2.felkido else 2)
                                while(beosztas_ml[h][bizhely1.bizoszlop].value is not None or beosztas_ml[utana][bizhely2.bizoszlop].value is not None): # and utana <= vizsgazokNapiMax):
                                    h += 1
                                    utana += 1
                                if(utana <= vizsgazokNapiMax):
                                    helykombinaciok_2.append([VizsgaIdopont(bizhely1.bizoszlop, beosztas_ml[1][bizhely1.bizoszlop].value, bizottsagok.vanFelkido(bizhely1.bizoszlop), h, napok[nap]),
                                                                VizsgaIdopont(bizhely2.bizoszlop, beosztas_ml[1][bizhely2.bizoszlop].value, bizottsagok.vanFelkido(bizhely2.bizoszlop), utana, napok[nap])])

            helykombinaciok_2.sort(key=lambda x: x[0].idopont)

            if(helykombinaciok_2[0][1].idopont > vizsgazokNapiMax):
                diakok_holnap.lista.append(d)
            else:
                beosztas_ml[ helykombinaciok_2[0][0].idopont ][ helykombinaciok_2[0][0].bizoszlop ].value = d.om
                diakok.ujVizsgaidopont(d.om, helykombinaciok_2[0][0].bizoszlop, beosztas_ml[1][helykombinaciok_2[0][0].bizoszlop].value, bizottsagok.vanFelkido(helykombinaciok_2[0][0].bizoszlop), helykombinaciok_2[0][0].idopont, napok[nap])

                beosztas_ml[ helykombinaciok_2[0][1].idopont ][ helykombinaciok_2[0][1].bizoszlop ].value = d.om
                diakok.ujVizsgaidopont(d.om, helykombinaciok_2[0][1].bizoszlop, beosztas_ml[1][helykombinaciok_2[0][1].bizoszlop].value, bizottsagok.vanFelkido(helykombinaciok_2[0][1].bizoszlop), helykombinaciok_2[0][1].idopont, napok[nap])

        # 3 vagy több osztályba jelentkezők ---------------------------------------------------------------------
        elif( len(d.tantargyak) >= 3 ):

            
            if(len(d.tantargyak) > 3):
                d.tantargyak.sort(key=lambda x: tantargyak_terheltsege[tantargyak.index(x)])

            helykombinaciok_3 = []

            for k in [[0,1,2], [0,2,1], [1,0,2], [1,2,0], [2,0,1], [2,1,0]]:

                bizottsag1 = next(biz for biz in bizottsagok.lista if biz.tantargy == d.tantargyak[k[0]])
                bizottsag2 = next(biz for biz in bizottsagok.lista if biz.tantargy == d.tantargyak[k[1]])
                bizottsag3 = next(biz for biz in bizottsagok.lista if biz.tantargy == d.tantargyak[k[2]])

                if bizottsag3.felkido or True not in [bizottsag1.felkido, bizottsag2.felkido, bizottsag3.felkido]:

                    set1 = set({})
                    for x in bizottsag1.bizhelyek:
                        set1.add(x.epresz)

                    set2 = set({})
                    for x in bizottsag2.bizhelyek:
                        set2.add(x.epresz)

                    set3 = set({})
                    for x in bizottsag3.bizhelyek:
                        set3.add(x.epresz)
                    
                    nincs_egy_szinten = False if len(set1 & set2 & set3)>0 else True

                    for bizhely1 in bizottsag1.bizhelyek:
                        for bizhely2 in bizottsag2.bizhelyek:
                            for bizhely3 in bizottsag3.bizhelyek:
                                if(bizhely1.epresz == bizhely2.epresz == bizhely3.epresz or nincs_egy_szinten):
                                    h = helykereso_1(bizhely1.bizoszlop, d.legkorabban)
                                    if(h > 0):
                                        elotte2 = h-(5 if bizottsag1.felkido else 2)
                                        elotte3 = elotte2-(5 if bizottsag2.felkido else 2)
                                        while( elotte3<d.legkorabban+5 or 
                                                (beosztas_ml[h][bizhely1.bizoszlop].value is not None or
                                                beosztas_ml[elotte2][bizhely2.bizoszlop].value is not None or
                                                beosztas_ml[elotte3][bizhely3.bizoszlop].value is not None) ):
                                            h += 1
                                            elotte2 += 1
                                            elotte3 += 1

                                        if(elotte3>=d.legkorabban+5):
                                            helykombinaciok_3.append([VizsgaIdopont(bizhely3.bizoszlop, beosztas_ml[1][bizhely3.bizoszlop].value, bizottsagok.vanFelkido(bizhely3.bizoszlop), elotte3, napok[nap]),
                                                                    VizsgaIdopont(bizhely2.bizoszlop, beosztas_ml[1][bizhely2.bizoszlop].value, bizottsagok.vanFelkido(bizhely2.bizoszlop), elotte2, napok[nap]),
                                                                    VizsgaIdopont(bizhely1.bizoszlop, beosztas_ml[1][bizhely1.bizoszlop].value, bizottsagok.vanFelkido(bizhely1.bizoszlop), h, napok[nap])])


            helykombinaciok_3.sort(key=lambda x: x[0].idopont)

            if(helykombinaciok_3[0][2].idopont > vizsgazokNapiMax):
                diakok_holnap.lista.append(d)
            else:
                beosztas_ml[ helykombinaciok_3[0][0].idopont ][ helykombinaciok_3[0][0].bizoszlop ].value = d.om
                diakok.ujVizsgaidopont(d.om, helykombinaciok_3[0][0].bizoszlop, beosztas_ml[1][helykombinaciok_3[0][0].bizoszlop].value, bizottsagok.vanFelkido(helykombinaciok_3[0][0].bizoszlop), helykombinaciok_3[0][0].idopont, napok[nap])

                beosztas_ml[ helykombinaciok_3[0][1].idopont ][ helykombinaciok_3[0][1].bizoszlop ].value = d.om
                diakok.ujVizsgaidopont(d.om, helykombinaciok_3[0][1].bizoszlop, beosztas_ml[1][helykombinaciok_3[0][1].bizoszlop].value, bizottsagok.vanFelkido(helykombinaciok_3[0][1].bizoszlop), helykombinaciok_3[0][1].idopont, napok[nap])

                beosztas_ml[ helykombinaciok_3[0][2].idopont ][ helykombinaciok_3[0][2].bizoszlop ].value = d.om
                diakok.ujVizsgaidopont(d.om, helykombinaciok_3[0][2].bizoszlop, beosztas_ml[1][helykombinaciok_3[0][2].bizoszlop].value, bizottsagok.vanFelkido(helykombinaciok_3[0][2].bizoszlop), helykombinaciok_3[0][2].idopont, napok[nap])

                if len(d.tantargyak) > 3:
                    d.tantargyak.remove(d.tantargyak[0])
                    d.tantargyak.remove(d.tantargyak[0])
                    d.tantargyak.remove(d.tantargyak[0])
                    diakok_holnap.lista.append(d)

    print("")

    # Eredmény kiírása

    # print(napok[nap])
    # print("")
    # for x in beosztas_ml.rows:
    #     for y in x:
    #         if y.value is not None:
    #             print(f"{str(y.value):11}", end=" ")
    #         else:
    #             print(f"{' ':11}", end=" ")
    #     print("")

    if(len(diakok_holnap.lista)>0):
        diakok_akt.lista = diakok_holnap.lista
        diakok_holnap.lista = []
        nap += 1
    else:
        tovabb = False

#print("")

#////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////
# Egy nap beosztása - VEGE
#////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////


#===============================================================================================
# Kimenetek előállítása
#
#------------------------------------------------------
#

#locale.setlocale(locale.LC_COLLATE, ('hu_HU', 'utf-8'))

diakok.lista.sort(key=lambda x: strkonv(x.nev))


# Vizsgázók listázása naponként
#
print("Vizsgázók listázása naponként")
maxNap = nap
nap = 0

while(nap <= maxNap):

    beosztas_ml = adatok_tbl[napok[nap]]

    print(napok[nap])

    adatok_tbl.create_sheet(napok[nap]+" beosztás")
    beosztas2_ml = adatok_tbl[napok[nap]+" beosztás"]

    sor = 1

    for diak in diakok.lista:
        if(napok[nap] in [x.nap for x in diak.vizsgaIdopontok]):
            if(len(diak.tantargyak)>0):
                #print(diak.nev, end=": ")
                beosztas2_ml.cell(row=sor, column=1).value = diak.nev
                oszlop = 2
                for x in diak.vizsgaIdopontok:
                    if(napok[nap] == x.nap):
                        if(x.felkido):
                            if((x.idopont < 11) and (x == diak.vizsgaIdopontok[0])):
                                #print(x.nap, str(beosztas_ml[5][0].value)[:5], "-", beosztas_ml[2][x.bizoszlop].value, "-", beosztas_ml[1][x.bizoszlop].value, end=" ")
                                beosztas2_ml.cell(row=sor, column=oszlop).value = x.nap + " " + str(beosztas_ml[5][0].value)[:5] + " - " + str(beosztas_ml[2][x.bizoszlop].value) + " - " + str(beosztas_ml[1][x.bizoszlop].value)
                            else:
                                #print(x.nap, str(beosztas_ml[x.idopont-3][0].value)[:5], "-", beosztas_ml[2][x.bizoszlop].value, "-", beosztas_ml[1][x.bizoszlop].value, end=" ")
                                beosztas2_ml.cell(row=sor, column=oszlop).value = x.nap + " " + str(beosztas_ml[x.idopont-3][0].value)[:5] + " - " + str(beosztas_ml[2][x.bizoszlop].value) + " - " + str(beosztas_ml[1][x.bizoszlop].value)
                        else:
                            #print(x.nap, str(beosztas_ml[x.idopont][0].value)[:5], "-", beosztas_ml[2][x.bizoszlop].value, "-", beosztas_ml[1][x.bizoszlop].value, end=" ")
                            beosztas2_ml.cell(row=sor, column=oszlop).value = x.nap + " " + str(beosztas_ml[x.idopont][0].value)[:5] + " - " + str(beosztas_ml[2][x.bizoszlop].value) + " - " + str(beosztas_ml[1][x.bizoszlop].value)
                        oszlop += 1
    
            #print("")   
            sor += 1

    for col in range(beosztas2_ml.max_column):
        beosztas2_ml.column_dimensions[chr(65+col)].width = 34

    nap += 1


# Vizsgabizottságok időbeosztása
#
print("Vizsgabizottságok időbeosztása")
nap = 0

while(nap <= maxNap):

    beosztas_ml = adatok_tbl[napok[nap]]

    bizottsagok = Bizottsagok()
    for tantargy in enumerate(beosztas_ml[4]):
        if tantargy[1].value != "Tantárgyak:":
            bizottsagok.ujBizottsag(tantargy[1].value, tantargy[0], True if type(beosztas_ml[5][tantargy[0]].value) != int else False,
                                        beosztas_ml[3][tantargy[0]].value, beosztas_ml[2][tantargy[0]].value)
    print(f"{napok[nap]}: ", end="")
    for tant in bizottsagok.lista:
        print(tant.tantargy, end=" - ")
        
        for biz in tant.bizhelyek:

            adatok_tbl.create_sheet(napok[nap]+" "+beosztas_ml[1][biz.bizoszlop].value)
            beosztas2_ml = adatok_tbl[napok[nap]+" "+beosztas_ml[1][biz.bizoszlop].value]

            beosztas2_ml[1][0].value = beosztas_ml[1][biz.bizoszlop].value
            beosztas2_ml[2][0].value = beosztas_ml[2][biz.bizoszlop].value
            beosztas2_ml[3][0].value = napok[nap]

            if tant.felkido:
                beosztas2_ml.cell(row = 5, column = 1).value = "Bejön"
                beosztas2_ml.cell(row = 5, column = 2).value = "Név"
                beosztas2_ml.cell(row = 5, column = 3).value = "Felel"
            else:
                beosztas2_ml.cell(row = 5, column = 1).value = "Idő"
                beosztas2_ml.cell(row = 5, column = 2).value = "Név"

            sor = 6
            while(beosztas_ml[sor-1][0].value is not None and beosztas_ml[sor+(2 if tant.felkido else -1)][biz.bizoszlop].value is not None):
                if tant.felkido:
                    if sor < 9:
                        beosztas2_ml.cell(row=sor, column=1).value = beosztas_ml[5][0].value
                    else:
                        beosztas2_ml.cell(row=sor, column=1).value = beosztas_ml[sor-1][0].value
                    beosztas2_ml.cell(row=sor, column=2).value = next(x.nev for x in diakok.lista if x.om == beosztas_ml[sor+2][biz.bizoszlop].value)
                    beosztas2_ml.cell(row=sor, column=3).value = beosztas_ml[sor+2][0].value
                else:
                    beosztas2_ml.cell(row=sor, column=1).value = beosztas_ml[sor-1][0].value
                    beosztas2_ml.cell(row=sor, column=1).number_format = "h:mm"
                    beosztas2_ml.cell(row=sor, column=2).value = next(x.nev for x in diakok.lista if x.om == beosztas_ml[sor-1][biz.bizoszlop].value)
                sor += 1

    nap += 1
    print("")

# locale.setlocale(locale.LC_ALL, "")

adatok_tbl.template = False
adatok_tbl.save('teszt.xlsx')